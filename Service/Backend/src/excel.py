from io import BytesIO
from typing import List, Self, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class BaseExcel:
    workbook: Workbook
    last_sheet: str
    fields: List[str]

    def __init__(self):
        self.workbook = Workbook()
        self.worksheets = []

        # 移除預設 Sheet
        self.workbook.remove(self.workbook.active)

    def add_worksheet(self, sheet_name: str):
        self.workbook.create_sheet(title=sheet_name)
        self.last_sheet = sheet_name
        self.fields = []

    def set_header_fields(self: Self, fields: List[str]=[]):
        self.fields = fields
        current_sheet = self.workbook[self.last_sheet]

        # 加入欄位名稱
        for c_idx, col_name in enumerate(fields, start=1):
            cell = current_sheet.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 預設欄寬
        for col_idx in range(1, current_sheet.max_column + 1):
            col_letter = current_sheet.cell(row=1, column=col_idx).column_letter
            current_sheet.column_dimensions[col_letter].width = 12

    def add_style_field(self, field, **kwargs):
        self.fields.append(kwargs.get("title"))
        current_sheet = self.workbook[self.last_sheet]
        c_idx = len(self.fields)

        # 預設 Style
        cell_title = field
        cell_width = kwargs.get("width") if kwargs.get("width") else 12
        cell_font_color = kwargs.get("font_color") if kwargs.get("font_color") else "FFFFFF"
        cell_bg_color = kwargs.get("bg_color") if kwargs.get("bg_color") else "009100"

        cell = current_sheet.cell(row=1, column=c_idx, value=cell_title)
        cell.font = Font(bold=True, color=cell_font_color)
        cell.fill = PatternFill("solid", fgColor=cell_bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_sheet.column_dimensions[cell.column_letter].width = cell_width

    def add_row_data(self, info: List[Any], style: dict={}):
        current_sheet = self.workbook[self.last_sheet]
        r_idx = current_sheet.max_row + 1

        cell_bg_color = style.get("bg_color") if style.get("bg_color") else "FFFFFF"

        for (d_idx, d_val) in enumerate(info):
            cell = current_sheet.cell(row=r_idx, column=d_idx + 1, value=d_val)
            cell.fill = PatternFill("solid", fgColor=cell_bg_color)


    def add_spare_row(self, color: str="ADADAD"):
        current_sheet = self.workbook[self.last_sheet]
        r_idx = current_sheet.max_row + 1
        fields_count = len(self.fields)
        current_sheet.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=fields_count)

        cell = current_sheet.cell(row=r_idx, column=1, value="--- 跨日線 ---")
        cell.fill = PatternFill("solid", fgColor=color)

    @property
    def steam(self: Self) -> BytesIO:
        for sheet_name in self.workbook.sheetnames:
            curr_sht = self.workbook[sheet_name]

            for row in curr_sht.iter_rows(min_row=1, max_row=curr_sht.max_row, max_col=curr_sht.max_column):
                for cell in row:
                    # 加上邊框 & 對齊
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # 儲存到記憶體
        stream = BytesIO()
        self.workbook.save(stream)
        stream.seek(0)

        return stream
